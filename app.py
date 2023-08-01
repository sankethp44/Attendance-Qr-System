import logging
import traceback
from azure.cosmosdb.table.tableservice import TableService
from email.message import EmailMessage
from flask import Flask, render_template, request, redirect, session, jsonify, send_file
import json
from flask import Flask, render_template, request, redirect, session, send_file
from azure.storage.blob import BlobServiceClient
import openpyxl,io
from azure.common import AzureMissingResourceHttpError
from urllib.parse import quote
import openpyxl
from azure.cosmosdb.table.tableservice import TableService
from azure.cosmosdb.table.models import Entity
import requests
from azure.common import AzureMissingResourceHttpError
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
from io import BytesIO
from email.mime.image import MIMEImage
from azure.storage.blob import BlobServiceClien

app = Flask(__name__)
app.secret_key = os.environ.get('app.secret_key')
# Replacing with actual Azure Storage account credentials
storage_account_name = 'blobdatabase234'
storage_account_key = os.environ.get('storage_account_key')
container_name = 'login'
excel_file_name = 'credentials.xlsx'
excel_file_name1 = 'studentform.xlsx'
TABLE_NAME = "Records"
connection_stringkey = os.environ.get('connection_stringkey')
table_service = TableService(account_name=storage_account_name, account_key=storage_account_key)

def save_student_data_to_excel(name, roll_no, email):
    # Replacing with Azure Table storage account connection string
    table_service = TableService(connection_string=connection_stringkey)

    # Define the name of Azure Table
    table_name = "studentdata1"

    # Create the table if it doesn't exist
    if not table_service.exists(table_name):
        table_service.create_table(table_name)

    # Create a new entity and set its properties (columns)
    student_entity = {
        "PartitionKey": name,
        "RowKey": roll_no,
        "Email": email,
    }

    # Inserting the entity into the Azure Table
    table_service.insert_or_replace_entity(table_name, student_entity)

def save_student_data_to_excel1(name, roll_no, email, status):
    # Replacing with Azure Table storage account connection string
    table_service = TableService(connection_string=connection_stringkey)

    # Defining the name of Azure Table
    table_name = "Records"

    # Create the table if it doesn't exist
    if not table_service.exists(table_name):
        table_service.create_table(table_name)

    # Create a new entity and set its properties (columns)
    student_entity = {
        "PartitionKey": name,
        "RowKey": roll_no,
        "Email": email,
        "Status": status,
    }
    # Inserting the entity into the Azure Table
    table_service.insert_or_replace_entity(table_name, student_entity)


@app.route('/delete_all_rows', methods=['POST'])
def delete_all_rows():
    # Connecting to the Azure Table storage
    table_service = TableService(connection_string=connection_stringkey)

    # Geting all entities from the table
    entities = table_service.query_entities('studentdata1')

    # Delete each entity (row) from the table
    for entity in entities:
        table_service.delete_entity('studentdata1', entity.PartitionKey, entity.RowKey)

    entities = table_service.query_entities('Records')
    # Delete each entity (row) from the table
    for entity in entities:
        table_service.delete_entity('Records', entity.PartitionKey, entity.RowKey)

    # Replacing with the name of your container containing the QR codes
    container_name = "qrcodes"

    # Initialize the BlobServiceClient using the connection string
    blob_service_client = BlobServiceClient.from_connection_string(connection_stringkey)

    # Getting a list of blobs in the "qrcode" container
    container_client = blob_service_client.get_container_client(container_name)
    blobs = container_client.list_blobs()

    # Deleting each blob from the "qrcodes" container
    for blob in blobs:
        container_client.delete_blob(blob)


    return render_template('form.html')


@app.route('/', methods=['GET', 'POST'])
def login():
    error_message = None  
    if request.method == 'POST':
        # Retrieve login form data
        username = request.form['username']
        password = request.form['password']

        # Connecting to Azure Storage Blob
        container_name = 'login'
        
        blob_service_client = BlobServiceClient.from_connection_string(connection_stringkey)
        container_client = blob_service_client.get_container_client(container_name)

        # Authenticating user from the Excel file
        authenticated = authenticate_user(username, password, blob_service_client, container_name, excel_file_name)

        if authenticated:
            return redirect('/homepage')
        else:
            error_message = "Please check your credentials, sir/madam please type username as frt and password as frt(This is for demo purpose only)"  # Set error message

    return render_template('index.html', error_message=error_message)  # Pass error_message to template


@app.route('/homepage', methods=['GET', 'POST'])
def HomePage():
    return render_template('HomePage.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')


@app.route('/about')
def about():
    return render_template('about.html')

def authenticate_user(username, password, blob_service_client, container_name, excel_file_name):
    # Getting the blob client for the credentials.xlsx file in the login container
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=excel_file_name)

    try:
        # Downloading the credentials.xlsx file from Azure Blob storage
        blob_data = blob_client.download_blob()
        blob_content = blob_data.readall()

        # Loading the Excel file from the blob content using BytesIO
        excel_file = BytesIO(blob_content)
        workbook = openpyxl.load_workbook(excel_file)

        sheet = workbook.active

        # Iterating through the rows and compare credentials
        for row in sheet.iter_rows(values_only=True):
            if row[0] == username and row[1] == password:
                return True

    except Exception as e:
        print(f"Error: {e}")

    return False

@app.route('/form')
def student_form():
    return render_template('form.html')


@app.route('/submit_form', methods=['GET', 'POST'])
def submit_form():
    if request.method == 'POST':
        # Retrieving form data from the request
        name = request.form['name']
        roll_no = request.form['roll_no']
        email = request.form['email']

        # Saving the data to the Excel file in Azure Blob Storage
        save_student_data_to_excel(name, roll_no, email)

        submitted = True

        return render_template('form.html',submitted=submitted)
    
#Start of email code
# Function to fetch student data and QR code URLs from Azure Table storage
def fetch_student_data_from_table():
    # Replacing with your Azure Table storage account connection string
    table_service = TableService(connection_string=connection_stringkey)

    # Replacing with your Azure Table name
    table_name = "studentdata1"

    # Fetching all entities from the Azure Table
    entities = table_service.query_entities(table_name)

    # Initializing a list to store student data and QR code URLs
    qr_codes_data = []

    # Iterating through each entity and get the student data and QR code URLs
    for entity in entities:
        name = entity.get('PartitionKey', '')
        roll_no = entity.get('RowKey', '')
        email = entity.get('Email', '')
        qr_codes_data.append({"name": name, "roll_no": roll_no, "email": email})

    return qr_codes_data
# Functioning to send an email with an attachment


# Flask route to fetch QR codes and display success message
@app.route('/generate_qr_codes', methods=['POST'])
def generate_qr_codes():
    # Making a GET request to the Azure Function endpoint that generates QR codes
    response = requests.get("https://qrcode111.azurewebsites.net/api/generateallqrcodes?")  # Replacing with the URL of your Azure Function
    # Checking the status code to determine if the request was successful
    if response.status_code == 200:
        # Rendering the form.html template with the success message
        return render_template('form.html')
    else:
        # Handling the case where the request was not successful
        return render_template('form.html')


def send_email(smtp_username, smtp_password, sender_email, receiver_email, subject, message, qr_code_urls):
    # Seting up the MIMEMultipart object with the email content
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    # Attaching the plain text message to the email
    msg.attach(MIMEText(message, 'plain'))

    # Fetching QR code images from Blob storage and attach them to the email
    for qr_code_url in qr_code_urls:
        # Fetching the QR code image data from the Blob storage URL
        response = requests.get(qr_code_url)
        if response.status_code == 200:
            qr_code_data = response.content

            # Extracting the filename from the URL and encode it properly
            qr_code_filename = qr_code_url.split("/")[-1]

            # Attaching the QR code image to the email
            qr_code_attachment = MIMEImage(qr_code_data, name=qr_code_filename)
            msg.attach(qr_code_attachment)

    try:
        # Connecting to the SMTP server and send the email
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)

        logging.info(f"Email sent successfully to {receiver_email}")
        return True
    except Exception as e:
        logging.error(f"Failed to send email to {receiver_email}. Error: {e}")
        logging.error(traceback.format_exc())  # Print the full traceback for detailed error analysis
        return False

#####


@app.route('/sendmail', methods=['POST'])
def sendmail():
    # Replacing with the SMTP email credentials
    smtp_username = 'spalaksha@gmail.com'
    smtp_password = os.environ.get('smtp_password')
    sender_email = 'spalaksha@gmail.com'

    # Fetching student data and QR code URLs from Azure Table storage
    qr_codes_data = fetch_student_data_from_table()

    # Fetching the Blob Service Client
    blob_service_client = BlobServiceClient.from_connection_string(connection_stringkey)
    container_client = blob_service_client.get_container_client("qrcodes")

    # Looping through the QR codes data and send emails to each student
    for student_info in qr_codes_data:
        name = student_info['name']
        roll_no = student_info['roll_no']
        email = student_info['email']
        message = f"Dear {name},\n\nPlease download this QR code to receive attendance for tomorrow's Annual Day event.\n\nBest Regards,\nYour friend"

        # Geting the URL of the QR code image from Azure Blob Storage
        qr_code_url = container_client.get_blob_client(f"{name}_{roll_no}.png").url
        # Sending the email with the QR code URL as an attachment
        send_email(smtp_username, smtp_password, sender_email, email, "Attendance QR Code", message, [qr_code_url])
    # Rendering the form.html template with the success message
    return render_template('form.html', message="Generated QR Codes Successfully! Check your email for the QR codes.")

#End of email



# Function to fetch student data from Azure Table and save it to Excel

def save_student_data_to_excel2():
    # Replacing with the Azure Table storage account connection string
    table_service = TableService(connection_string=connection_stringkey)

    # Defining the name of your Azure Table
    table_name = "Records"

    # Querying all entities from the Azure Table
    entities = table_service.query_entities(table_name)

    # Creating an Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Writing the headers
    headers = ["Name", "Roll No", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)

    # Writing the student data to the worksheet
    row_num = 2
    for student in entities:
        student_data = [student.PartitionKey, student.RowKey, student.Status]
        for col_num, data in enumerate(student_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=data)
        row_num += 1

    # Saving the Excel workbook to a BytesIO buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    # Returning the Excel data as a BytesIO object
    return excel_buffer

# Flask route to handle the download of the Records table
@app.route('/download_records')
def download_records():
    # Geting the Excel data from the "Records" table in Azure
    excel_data = save_student_data_to_excel2()
    if excel_data is None:
        # If the table is empty, provide a message saying Table is empty
        return "The table is empty. No records to display."
    # Sending the Excel data as a downloadable file to the user
    return send_file(
        excel_data,
        as_attachment=True,
        download_name='Records.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Home Page
@app.route('/')
def home_page():
    return render_template('HomePage.html')

# Video Stream Page
@app.route('/video_stream')
def video_stream_page():
    return render_template('video_stream.html')


@app.route('/scanner', methods=['POST'])
def handle_scanner_data():
    if request.method == 'POST':
        qr_data = request.form['data']
        name, roll_no = qr_data.split(':')

        # Saving the student data to the Azure Table
        save_student_data_to_table(name, roll_no)

        # Showing confirmation message using JavaScript alert
        return render_template('video_stream.html')

    # Handling other HTTP methods (if needed)
    return "Invalid request method.", 405

def is_duplicate(name, roll_no):
    # Initializing the TableService with the storage account name and key
    table_service = TableService(account_name=storage_account_name, account_key=storage_account_key)

    # Querying the table to check if the roll_no already exists
    query_filter = f"PartitionKey eq '{name}' and RowKey eq '{roll_no}'"
    entities = table_service.query_entities(TABLE_NAME, filter=query_filter)

    # If any entities are found, then it is a duplicate QR code
    return any(entities)

def save_student_data_to_table(name, roll_no):
    # Initializing the TableService with the storage account name and key
    table_service = TableService(account_name=storage_account_name, account_key=storage_account_key)

    # Checking if the entity with the given name and roll_no exists
    try:
        existing_entity = table_service.get_entity(TABLE_NAME, name, roll_no)
    except AzureMissingResourceHttpError as ex:
        # If the entity does not exist, insert a new entity
        student_data = {
            'PartitionKey': name,
            'RowKey': roll_no,
            'Status': 'Present'  # You can set the Status to 'Present'
        }
        table_service.insert_entity(TABLE_NAME, student_data)
        return

    # If the entity exists, update its Status to 'Present'
    existing_entity['Status'] = 'Present'
    table_service.update_entity(TABLE_NAME, existing_entity)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)